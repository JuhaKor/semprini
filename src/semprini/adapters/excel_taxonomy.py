"""Bundled Excel taxonomy adapter (spec 5.3).

One workbook is one taxonomy, and one taxonomy is one configured source: the scheme's
metadata comes from the workbook's own ``Concept Scheme`` sheet, and every value in it
carries a source ref under that source's name. Adding a second taxonomy is a second
source entry, which is what makes ``sem:sourceRef`` say *which file* an object came from
without the adapter's name being frozen into identity (spec 5.4).

Hierarchy is **ragged**: depth is a value's position across ``L1..Ln`` preferred-label
columns rather than a ``parent_code`` pointing at another row. A value at depth *k* is
narrower than the row whose path is its first *k-1* labels. Two consequences fall out of
that shape and are worth stating, because they are not obvious:

*A cycle cannot be expressed.* A row's ancestors are a prefix of its own cells, so there
is nothing to close a loop with. The three error conditions are dangling parents,
duplicates and skipped levels instead.

*A label is structural.* Renaming an L2 cell re-parents everything beneath it. That is
why identity comes from the ``Concept URI`` column and never from the labels: the ID map
is keyed by that identifier (spec 5.4), so a taxonomy can be re-worded without minting a
single new IRI.
"""

from __future__ import annotations

import re
from collections.abc import Iterator, Mapping, Sequence
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from semprini.adapters.base import AdapterError, BaseAdapter, SourceUnreachableError
from semprini.config import ConfigError, escapes_the_instance, is_slug
from semprini.model import (
    InternalModel,
    Issue,
    Scheme,
    SchemeType,
    Severity,
    SourceRef,
    TaxonomyValue,
    Text,
    is_language_tag,
)

__all__ = ["ExcelTaxonomyAdapter"]

SCHEME_SHEET = "Concept Scheme"
TAXONOMY_SHEET = "Taxonomy"

_SETTINGS = frozenset({"path", "scheme_slug", "enumerates_source"})

# Scheme-sheet rows this adapter reads. Every other row in the sheet — creator, dates,
# version, domain, the scheme URI and prefix — is documentation for whoever maintains the
# workbook. Reading them would mean deciding where each lands in the metamodel, and three
# of them are hand-typed dates in a spreadsheet.
_SCHEME_NAME = "scheme name"
_SCHEME_DESCRIPTION = "description"
_SCHEME_LANGUAGE = "language"
_SCHEME_ENUMERATES = "reference entity uuid"

_CONCEPT_URI = "concept uri"
_DEFINITION = "definition"
_ALT_LABELS = "alternative labels"
_HIDDEN_LABELS = "hidden labels"
_SCOPE_NOTE = "scope note"
_EXAMPLE = "example"

# The value columns carried into the model, beside the identity and level columns.
# Anything else in the sheet is tolerated and ignored — a workbook is a working document
# and gains columns for reasons of its own, so an unknown header is not an error the way
# an unknown *config* key is (spec 5.1). Notes and the provenance columns fall here: they
# have no home in the metamodel, and a hand-typed extraction date does not belong in a
# graph the compiler regenerates.
_VALUE_COLUMNS = (_DEFINITION, _ALT_LABELS, _HIDDEN_LABELS, _SCOPE_NOTE, _EXAMPLE)

_LEVEL_HEADER = re.compile(r"^l(\d+)\s*-\s*preferred label")

# A cell written in Turtle's literal syntax: "Computers & Tablets"@en. The workbook states
# languages per cell, so both branches of spec 5.5 rule 6 are reachable through this
# adapter — a tagged cell keeps its tag, and a bare one takes the scheme's language or,
# failing that, the instance's.
_LITERAL = re.compile(r'^\s*"([^"]*)"(?:@([A-Za-z0-9-]+))?\s*$', re.DOTALL)


class TaxonomyContentError(AdapterError):
    """A workbook was read but says something the compiler cannot act on.

    Exit code 1, not 3: the file was perfectly reachable and its content is wrong, which
    is a steward's problem rather than a retry (spec 5.1).

    Every problem in the workbook is collected into one of these rather than raised at the
    first. A taxonomy is edited in bulk — a re-export, a re-levelled branch — so the
    mistakes come in bulk too, and reporting one per CI run costs an operator a round trip
    each. That is a deliberate departure from the shape of the other adapter errors, which
    are single-cause by nature.
    """

    def __init__(self, path: Path, issues: Sequence[Issue]) -> None:
        self.issues = tuple(issues)
        listed = "\n".join(f"  - {issue}" for issue in self.issues)
        count = f"{len(self.issues)} problem" + ("s" if len(self.issues) != 1 else "")
        super().__init__(f"{path}: {count}\n{listed}")


class ExcelTaxonomyAdapter(BaseAdapter):
    """A taxonomy workbook: one file, one concept scheme, a ragged label hierarchy.

    The line above is what ``semprini adapters`` prints beside this adapter's name.

    Settings:

    ``path``
        The workbook, relative to the instance repository. Committed under
        ``sources/taxonomies/`` so that a taxonomy edit and the generated Turtle land in
        one reviewable PR (spec 4.2).
    ``scheme_slug``
        The scheme's permanent slug. In the configuration rather than the workbook
        because it names the scheme's IRI local name *and* its output file, and both are
        frozen by the ID map on the run that mints them (spec 3.4.2, 4.2).
    ``enumerates_source``
        The configured source that issued the ``Reference Entity UUID`` — the modelling
        tool, normally the instance's ``ellie`` source. Required exactly when that cell
        is filled, and in the configuration rather than the workbook because a source
        *name* is this instance's to choose (spec 5.1): the workbook states a UUID, and
        which configured source that UUID belongs to is not a fact about the workbook.
        Without it the reference would be looked up under the *taxonomy's* own source
        name, where an entity's key can never be found (spec 5.4).
    """

    name = "excel-taxonomy"

    _fetched: int | None = None
    """How many values the last fetch read, for :meth:`summary`. ``None`` until then — an
    adapter is constructed by ``semprini check`` without ever fetching (spec 6.1)."""

    # --------------------------------------------------------------------- the contract

    def fetch(self) -> InternalModel:
        # Its own configuration, checked before it is used. `validate_config()` is called
        # by `semprini check` (spec 6.1) and by the contract suite, and by nothing on the
        # compile path — so without this a run that skipped `check` would reach `_path()`
        # with a setting no one validated, where an absolute path silently wins over the
        # repository root and a missing key is a bare KeyError traceback rather than an
        # issue naming it. Exit 2, the same as any other configuration error.
        issues = [issue for issue in self.validate_config() if issue.severity is Severity.ERROR]
        if issues:
            raise ConfigError(issues)

        path = self._path()
        workbook = self._open(path)
        try:
            metadata = _read_scheme_sheet(self._sheet(workbook, SCHEME_SHEET, path), path)
            # The scheme's language decides how every cell below is read, so it is
            # validated before a single row is.
            language = metadata.get(_SCHEME_LANGUAGE)
            if language is not None and not is_language_tag(language):
                raise TaxonomyContentError(
                    path,
                    [
                        Issue(
                            Severity.ERROR,
                            f"not a language tag: {language!r}",
                            f"{SCHEME_SHEET}!Language",
                        )
                    ],
                )
            sheet = self._sheet(workbook, TAXONOMY_SHEET, path)
            rows = _read_taxonomy_sheet(sheet, path, language)
        finally:
            workbook.close()

        slug = str(self.config["scheme_slug"])
        values = _build_values(rows, path, source=self.source_name, slug=slug, language=language)
        self._fetched = len(values)
        return InternalModel(
            schemes=(self._scheme(metadata, slug, language),), taxonomy_values=values
        )

    def validate_config(self) -> list[Issue]:
        issues: list[Issue] = []
        where = f"sources.{self.source_name}.config"

        raw = self.config.get("path")
        if not raw:
            issues.append(
                Issue(Severity.ERROR, "a taxonomy source needs a 'path'", f"{where}.path")
            )
        elif escapes_the_instance(str(raw)):
            # The workbook is part of the instance and is reviewed with it; a path leading
            # out of the repository is content nobody reviewed (spec 4.2).
            issues.append(
                Issue(
                    Severity.ERROR,
                    f"path must be inside the instance repository, got {raw!r}",
                    f"{where}.path",
                )
            )

        slug = self.config.get("scheme_slug")
        if not slug:
            issues.append(
                Issue(
                    Severity.ERROR,
                    "a taxonomy source needs a 'scheme_slug'",
                    f"{where}.scheme_slug",
                )
            )
        elif not is_slug(str(slug)):
            # The same definition an instance id and a source name are held to: this
            # becomes an IRI local name and a file name, both permanent (spec 3.4.2).
            issues.append(
                Issue(
                    Severity.ERROR,
                    f"not a slug: {slug!r} (lower-case letters, digits, '-' and '_')",
                    f"{where}.scheme_slug",
                )
            )

        owner = self.config.get("enumerates_source")
        if owner is not None and not is_slug(str(owner)):
            # A source name, held to the definition config.py holds every source name to,
            # so that a typo is caught here rather than as an unresolvable reference two
            # stages later (spec 5.1).
            issues.append(
                Issue(
                    Severity.ERROR,
                    f"not a source name: {owner!r} (lower-case letters, digits, '-' and '_')",
                    f"{where}.enumerates_source",
                )
            )

        for key in sorted(set(self.config) - _SETTINGS):
            issues.append(Issue(Severity.ERROR, f"unknown setting {key!r}", f"{where}.{key}"))
        return issues

    def summary(self) -> str:
        if self._fetched is None:
            return ""
        return f"{self._fetched} values from {Path(str(self.config['path'])).name}"

    # ------------------------------------------------------------------------ internals

    def _scheme(self, metadata: Mapping[str, str], slug: str, language: str | None) -> Scheme:
        enumerates = metadata.get(_SCHEME_ENUMERATES)
        owner = str(self.config.get("enumerates_source") or "")
        if enumerates and not owner:
            # Exit 2, not a content error: the workbook is right and the configuration is
            # incomplete. Checked here rather than in validate_config() because only the
            # workbook knows whether the cell is filled, and a source that enumerates
            # nothing must not be made to configure something it does not use.
            raise ConfigError(
                [
                    Issue(
                        Severity.ERROR,
                        f"the workbook names {enumerates!r} as the entity this taxonomy "
                        f"enumerates, so 'enumerates_source' must name the configured "
                        f"source that issued that key — the modelling tool's source, not "
                        f"this one",
                        f"sources.{self.source_name}.config.enumerates_source",
                    )
                ]
            )
        return Scheme(
            # Keyed by the slug, not by the file name: the path lives in the
            # configuration precisely so a workbook can be moved or renamed without
            # re-keying every object in it (spec 5.4).
            source_refs={self.source_name: slug},
            pref_label=_text(metadata[_SCHEME_NAME], language),
            definition=_optional_text(metadata.get(_SCHEME_DESCRIPTION), language),
            slug=slug,
            scheme_type=SchemeType.TAXONOMY,
            # A key in the modelling tool, resolved against the ID map when the graph is
            # built (spec 3.3). Until that tool's source is configured and compiled there
            # is nothing to resolve, which is why the cell is optional — and the ref is
            # keyed under *that* source, not this one, since the ID map is keyed by
            # (source name, source key) and the entity's row carries the modelling tool's
            # name (spec 5.4).
            enumerates=SourceRef(owner, enumerates) if enumerates else None,
        )

    def _path(self) -> Path:
        return self.ctx.repo_root / str(self.config["path"])

    def _open(self, path: Path) -> Any:
        try:
            return load_workbook(path, data_only=True, read_only=False)
        except OSError as error:
            # The only failure that is exit 3: the workbook could not be read at all
            # (spec 5.1). Bad *content*, below, is a compile failure instead.
            raise SourceUnreachableError(
                f"source {self.source_name!r}: cannot read {path}: {error}"
            ) from error
        except Exception as error:  # openpyxl raises its own types for a corrupt file
            raise AdapterError(
                f"source {self.source_name!r}: {path} is not a readable workbook: {error}"
            ) from error

    def _sheet(self, workbook: Any, title: str, path: Path) -> Worksheet:
        if title not in workbook.sheetnames:
            raise TaxonomyContentError(
                path, [Issue(Severity.ERROR, f"the workbook has no {title!r} sheet")]
            )
        sheet: Worksheet = workbook[title]
        return sheet


# --------------------------------------------------------------------------- the sheets


def _normalize_header(value: object) -> str:
    """A header cell as a key: first line, lower-cased, stripped.

    Headers in these workbooks carry their SKOS mapping on a second line — ``Concept URI``
    then ``(local identifier)`` — which is documentation for whoever fills the sheet in
    and no part of the column's name.
    """
    if value is None:
        return ""
    return str(value).split("\n", 1)[0].strip().lower()


def _cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def _read_scheme_sheet(sheet: Worksheet, path: Path) -> Mapping[str, str]:
    """The vertical Property/Value table, as normalized property → value."""
    metadata = {
        _normalize_header(row[0]): _cell(row[1])
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True)
        if row and row[0] is not None
    }
    metadata = {key: value for key, value in metadata.items() if value}
    if not metadata.get(_SCHEME_NAME):
        raise TaxonomyContentError(
            path,
            [Issue(Severity.ERROR, "no 'Scheme Name' row", f"{SCHEME_SHEET}!A")],
        )
    return metadata


class _Row:
    """One taxonomy row, already split into the parts the model needs."""

    __slots__ = ("key", "label", "number", "path", "values")

    def __init__(
        self,
        number: int,
        key: str,
        path: tuple[str, ...],
        label: Text,
        values: Mapping[str, str],
    ):
        self.number = number
        self.key = key
        self.path = path
        """The label *values* of this row's level cells, ancestors first.

        Values rather than raw cells: the hierarchy is matched on these, and a workbook
        that writes ``"Tools"@en`` in one row and a bare ``Tools`` in another means the
        same branch both times. Matching raw cells would split it in two and report every
        row beneath the second spelling as an orphan."""

        self.label = label
        """This row's own preferred label — the deepest level cell, with its language."""

        self.values = values


def _read_taxonomy_sheet(sheet: Worksheet, path: Path, language: str | None) -> Sequence[_Row]:
    """Read the sheet into rows, refusing a header this adapter cannot act on.

    Header matching is strict, and deliberately so. A ragged format encodes its whole
    hierarchy in column *positions*, so a sheet whose level columns are named something
    else does not read as a broken taxonomy — it reads as an empty one, and an empty
    taxonomy compiles happily to a scheme with no values and deprecates everything that
    was in it. Failing on the header is the only place that mistake is visible.
    """
    header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    columns: dict[str, int] = {}
    levels: list[tuple[int, int]] = []
    for index, raw in enumerate(header):
        name = _normalize_header(raw)
        if not name:
            continue
        columns.setdefault(name, index)
        match = _LEVEL_HEADER.match(name)
        if match:
            levels.append((int(match.group(1)), index))
    levels.sort()

    found = ", ".join(sorted(columns)) or "no headers at all"
    if _CONCEPT_URI not in columns:
        raise TaxonomyContentError(
            path,
            [
                Issue(
                    Severity.ERROR,
                    f"the {TAXONOMY_SHEET!r} sheet has no 'Concept URI' column, which is "
                    f"where a value's permanent identity comes from; found {found}",
                    f"{TAXONOMY_SHEET}!1",
                )
            ],
        )
    if not levels:
        raise TaxonomyContentError(
            path,
            [
                Issue(
                    Severity.ERROR,
                    f"the {TAXONOMY_SHEET!r} sheet has no 'L1 - Preferred Label' column, so "
                    f"it states no hierarchy and no labels; found {found}",
                    f"{TAXONOMY_SHEET}!1",
                )
            ],
        )
    declared = [number for number, _ in levels]
    if declared != list(range(1, len(declared) + 1)):
        # Depth is read from a cell's position among the level columns, so the columns
        # themselves must be L1..Ln with nothing missing. A sheet starting at L2, or one
        # that lost its L3 in a re-export, is otherwise read as a *complete* hierarchy
        # with every value one level too shallow: the run succeeds, every skos:broader
        # moves, and the diff reads as a deliberate re-levelling nobody performed. This
        # is the same silent shift the skipped-level rule catches within a row, and the
        # two are only distinguishable here, at the header.
        raise TaxonomyContentError(
            path,
            [
                Issue(
                    Severity.ERROR,
                    f"the {TAXONOMY_SHEET!r} sheet's level columns are "
                    f"{', '.join(f'L{number}' for number in declared)}; they must run "
                    f"L1..L{len(declared)} with none missing, or every value's depth is "
                    f"read one level out",
                    f"{TAXONOMY_SHEET}!1",
                )
            ],
        )

    issues: list[Issue] = []
    rows: list[_Row] = []
    for number, raw_row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        level_columns = [index for _, index in levels]
        row = _read_row(raw_row, number, columns, level_columns, language, issues)
        if row is not None:
            rows.append(row)
    if issues:
        raise TaxonomyContentError(path, issues)
    return rows


def _read_row(
    raw: Sequence[object],
    number: int,
    columns: Mapping[str, int],
    level_columns: Sequence[int],
    language: str | None,
    issues: list[Issue],
) -> _Row | None:
    def value(name: str) -> str:
        index = columns.get(name)
        return _cell(raw[index]) if index is not None and index < len(raw) else ""

    cells = [_cell(raw[index]) if index < len(raw) else "" for index in level_columns]
    key = _local_name(value(_CONCEPT_URI))
    if not any(_cell(cell) for cell in raw):
        # A wholly blank row is spreadsheet punctuation, not a value. Judged across the
        # *whole* row rather than the identity and level cells alone: a row carrying a
        # definition but no identity yet is half-finished work, and dropping it silently
        # is how a value a steward believes they added never appears.
        return None
    where = f"{TAXONOMY_SHEET}!{number}"
    if not key:
        # The prototype skipped these silently, which for a taxonomy means a value
        # quietly disappearing from the instance on the next compile.
        issues.append(Issue(Severity.ERROR, "no 'Concept URI', so this row has no identity", where))
        return None

    filled = [index for index, cell in enumerate(cells) if cell]
    if not filled:
        issues.append(Issue(Severity.ERROR, f"{key!r} has no preferred label", where))
        return None
    if filled != list(range(len(filled))):
        # Depth is the *position* of the last filled cell, so the prototype's habit of
        # collecting non-empty cells and discarding their columns reads L1+L3 as depth 2
        # and silently attaches the value to the wrong parent.
        missing = ", ".join(f"L{index + 1}" for index in range(filled[-1]) if index not in filled)
        issues.append(Issue(Severity.ERROR, f"{key!r} skips a level: {missing} is empty", where))
        return None

    # Parsed here rather than at the point of use: the path is what the hierarchy is
    # matched on, and matching raw cells would make "Tools"@en and a bare Tools two
    # branches when the workbook means one.
    labels = [_text(cell, language) for cell in cells[: filled[-1] + 1]]
    values = {name: value(name) for name in _VALUE_COLUMNS}
    return _Row(number, key, tuple(text.value for text in labels), labels[-1], values)


def _local_name(reference: str) -> str:
    """``ont:Laptops`` → ``Laptops``.

    The prefix is a convenience for whoever writes the sheet and no part of the identity:
    the source key is what the ID map is keyed by (spec 5.4), and it should not change
    because a workbook adopted a different prefix.
    """
    return reference.rsplit(":", 1)[-1].strip()


# ------------------------------------------------------------------------ the model


def _build_values(
    rows: Sequence[_Row], path: Path, *, source: str, slug: str, language: str | None
) -> tuple[TaxonomyValue, ...]:
    """Turn rows into values, resolving the ragged hierarchy into ``skos:broader``."""
    issues: list[Issue] = []
    by_path: dict[tuple[str, ...], _Row] = {}
    by_key: dict[str, _Row] = {}

    for row in rows:
        where = f"{TAXONOMY_SHEET}!{row.number}"
        clash = by_key.get(row.key)
        if clash is not None:
            issues.append(
                Issue(
                    Severity.ERROR,
                    f"{row.key!r} is already the identity of row {clash.number}; one "
                    f"identifier is one value",
                    where,
                )
            )
        else:
            by_key[row.key] = row
        twin = by_path.get(row.path)
        if twin is not None:
            issues.append(
                Issue(
                    Severity.ERROR,
                    f"{' / '.join(row.path)!r} is already the path of row {twin.number}; "
                    f"two values cannot sit in one place in the hierarchy",
                    where,
                )
            )
        else:
            by_path[row.path] = row

    values: list[TaxonomyValue] = []
    for row in rows:
        parent: SourceRef | None = None
        if len(row.path) > 1:
            above = by_path.get(row.path[:-1])
            if above is None:
                issues.append(
                    Issue(
                        Severity.ERROR,
                        f"{row.key!r} is narrower than {' / '.join(row.path[:-1])!r}, which "
                        f"no row defines",
                        f"{TAXONOMY_SHEET}!{row.number}",
                    )
                )
                continue
            parent = SourceRef(source, above.key)
        # Written out rather than splatted from a dict: these are the reused SKOS fields
        # (spec 3.3), and a mapping from column name to field name would let a typo pass
        # the type checker and silently drop a column.
        scope_note = _optional_text(row.values[_SCOPE_NOTE], language)
        example = _optional_text(row.values[_EXAMPLE], language)
        values.append(
            TaxonomyValue(
                source_refs={source: row.key},
                pref_label=row.label,
                definition=_optional_text(row.values[_DEFINITION], language),
                schemes=(slug,),
                parent=parent,
                alt_labels=_texts(row.values[_ALT_LABELS], language),
                hidden_labels=_texts(row.values[_HIDDEN_LABELS], language),
                scope_notes=(scope_note,) if scope_note is not None else (),
                examples=(example,) if example is not None else (),
            )
        )

    if issues:
        raise TaxonomyContentError(path, issues)
    return tuple(values)


def _texts(raw: str, language: str | None) -> tuple[Text, ...]:
    """A semicolon-separated cell, as the sheet's own header says these are.

    Only the label columns are split. A scope note or an example is prose — the sample
    workbooks write ``Laptop, iPad, Desktop PC`` in one cell — and cutting it up on a
    punctuation mark would invent several statements where the source made one.
    """
    return tuple(_text(part, language) for part in _split(raw) if part)


def _split(raw: str) -> Iterator[str]:
    """Split on semicolons that are **outside** a quoted literal.

    Splitting first and parsing afterwards looks equivalent and is not: a workbook may
    write both ``"A; B"@fi; "C"@fi`` and mean two labels, and a naive split cuts the first
    one in half, leaving two fragments that still carry stray quotes and lose the ``@fi``
    the cell stated. The result is wrong RDF rather than an error, which is the worst
    available outcome for a governed file.
    """
    depth = 0
    current: list[str] = []
    for character in raw:
        if character == '"':
            depth = 1 - depth
        if character == ";" and depth == 0:
            yield "".join(current).strip()
            current = []
        else:
            current.append(character)
    yield "".join(current).strip()


def _text(raw: str, language: str | None) -> Text:
    """A cell as a text, honouring a language it states for itself (spec 5.5 rule 6).

    Three levels, narrowest first: a tag written into the cell, then the language the
    workbook declares for the whole scheme, then — by leaving it unstated here — the
    instance's ``default_language``, applied when the graph is built.

    A cell counts as literal syntax only when the quoted part holds no further quotation
    mark. Prose that merely opens and closes with one — ``"Smart" tools "here"`` — is a
    definition somebody wrote, not a literal, and a greedy match would silently delete its
    outer characters on the way into a governed file. The cost of the narrow rule is that
    a label genuinely containing a quotation mark keeps its outer quotes rather than
    losing them; taking a cell too literally is recoverable, and quietly editing it is not.
    """
    match = _LITERAL.match(raw)
    if match is None:
        return Text(raw, language)
    value, tag = match.group(1), match.group(2)
    return Text(value, tag or language)


def _optional_text(raw: str | None, language: str | None) -> Text | None:
    return _text(raw, language) if raw else None
